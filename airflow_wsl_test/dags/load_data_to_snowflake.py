import os, csv
from openpyxl import load_workbook  # to be tested that all required data in a real dataset is read properly

from airflow import DAG
from airflow.providers.snowflake.operators.snowflake import SnowflakeOperator
from airflow.operators.python_operator import PythonOperator
from airflow.utils.dates import days_ago


# Potential orchestration improvements to consider

    # - switch to LocalExecutor (and setup PostgreSQL for Airflow) if keeping it local or to CeleryExecutor if moving it to the cloud
    #   it's a must have, but running Airflow on Windows under WSL caused to many bugs even with LocalExecutor. I would use docker or cloud next time :)
    # - add dbt pipelines' triggers
    # - remove CSV production step. Possible option: direct transfer to warehouse tables with Python Snowpark reading .xlsx files from Showflake Stage
    # - optimize .xlsx reading performance / data set size limits by:
    #    (a) iterating via multiple files/folders - if we connect to more data that is stored in a partitioned datalake
    #    (b) splitting a file if it's too large, 
    #    (c) reading/writing line by line if a file is too large
    # - more data quality tests on real data, also for csv lib outputs
    # - see if copying data from stage / merging to warehouse makes more sense in DBT pipelines
    # - think of better "upsert" algorithm for big data (mb keep single id-column table to compare against). 
    #   OR general practice: focus on excluding possibilities for duplicate records to appear, it's cheaper then handle/remove them.
    # - simplify files paths functions
    # - split tables to different dags to run them separately
    # - add checks for stage > DW copying
    # - implement dynamic file names and mb folders for storing partitioned data-like structure in snowflake stage
    #   or store the datalake somewhere else


# Base directory for data on a local device
    # Ideally we should refer to a data lake in a production environment 
    # also with a DataLake we could load data to the warehouse directly, w/o copying to snowflake stage
file_directory = os.path.dirname(os.path.abspath(__file__))  # current file directory
base_directory = os.path.join(file_directory, '..', '..', 'sumup_data')  # base directory relatively to the file

# File to table mapping
table_names = [
    'device',
    'store',
    'transaction'
]

default_args = {
    'owner': 'admin',
    'email_on_failure': False,
    'email_on_retry': False,
    'depends_on_past': False,   # since we do one-time analysis that is manually triggered. 
                                # for an incrementally updated pipeline True would be better.
}

common_snowflake_args = {
    'snowflake_conn_id': 'snowflake_conn',  # connection is configured in Airflow web UI
    'database': 'SUMUP',
    'schema': 'RAW'
}

# FUNCTION TO CONVERT EXCEL TO CSV

    # Unfortunately Snowflake doesn't support SQL queries to .XLSX (unlike MS SQL Server)
    # Yet writing snowpark code in Python is not much different from using pandas libraries (which is prohibited by the task)

    # The function converts a file as a whole
    # largest file 'transaction.xlsx' has 1500 rows with 160KB = 0,11KB per row
    # 0.11kB per row * 100M rows = 11 GB >>> files up to 100M rows are loadable to memory 
    # (subject to the machine properties and computing power allocation)
    # assuming the real data source has files partitioned by year/month/day
    # we won't face even a 100M rows file and can process files as a whole
    # TBC topic - processing time and error handling for large files

def convert_excel_to_csv(excel_path,csv_path):

    excel_file = load_workbook(excel_path)               # loads the workbook and selects the first worksheet 
    data_sheet = excel_file.worksheets[0]               # selects first sheet in the excel file

    with open(csv_path, 'w', encoding='utf-8', newline='') as csv_file:   # Opens the CSV file
        writer = csv.writer(csv_file)
        
        for row in data_sheet.iter_rows():              # Iterates through the rows in the worksheet and write to the CSV file
            row_data = [cell.value for cell in row]
            writer.writerow(row_data)


def delete_file(file_path):  # isn't critical, should not break the pipeline
    try:
        os.remove(file_path)
        print(f"File {file_path} has been deleted.")
    except FileNotFoundError:
        print(f"File {file_path} not found.")
    except PermissionError:
        print(f"Permission denied for deleting {file_path}.")
    except Exception as e:
        print(f"An error occurred: {e}")


with DAG(
    'load_data_to_snowflake',
    default_args=default_args,
    description='Converts Excel files into CSVs, loads to Snowflake and ',
    schedule_interval=None,  # Manual trigger, since we consider 1-time analysis
    start_date=days_ago(1),
    catchup=False  # Since we consider 1-time analysis
) as dag:

    for table_name in table_names:

        csv_file_name = table_name + '.csv'
        excel_file_name = table_name + '.xlsx'
        csv_path = os.path.join(base_directory, csv_file_name)
        excel_path = os.path.join(base_directory, excel_file_name)  

        convert_to_csv = PythonOperator (
            task_id=f'convert_{excel_file_name}_to_csv',
            python_callable=convert_excel_to_csv,
            op_args=[excel_path,csv_path]
        )

        stage_sql = f'PUT file://{csv_path} @SUMUP_RAW_DATA AUTO_COMPRESS=TRUE'
        
        stage_task = SnowflakeOperator(
            task_id=f'stage_{csv_file_name}_to_snowflake',
            sql=stage_sql,
            **common_snowflake_args
        )

        copy_sql = f"""
            TRUNCATE {table_name};  -- [assignment only] Since we consider 1-time analysis, trancation helps to
                                    -- run the pipeline multiple times w/o manually cleaning databases
                                    
                                    -- For large datasets, "upsert" logic might be implemented for small tables to avoid duplicates
                                    -- when iterating through multiple files for the same table.
                                    -- for large tables (100M+ rows), "upsert" logic would be too heavy / or more sophisticated
            COPY INTO {table_name}
            FROM '@SUMUP_RAW_DATA/{os.path.basename(csv_path)}'
            FILE_FORMAT = (TYPE = 'CSV' FIELD_OPTIONALLY_ENCLOSED_BY = '"')
            ON_ERROR = 'CONTINUE';
            """

        copy_task = SnowflakeOperator(
            task_id=f'copy_{table_name}_to_warehouse',
            sql=copy_sql,
            **common_snowflake_args
        )

        delete_csv = PythonOperator (             # With large data and multiple files from a partitioned datalake
            task_id=f'delete_{file_name}.csv',    # we wouldn't want to keep both .xlsx and .csv copies
            python_callable=delete_file,          # opt for keeping the original file
            op_args=[csv_path]
        )


        # Task order
        convert_to_csv >> stage_task >> copy_task >> delete_csv  # csv deleted last to save compute resources if we need to re-run the pipeline